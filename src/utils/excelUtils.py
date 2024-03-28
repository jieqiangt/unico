from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import CellIsRule

def get_column_max_width(column):

    max_length = 0

    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    return (max_length + 2) * 1.2


def auto_adjust_column(worksheet, column, header_row=1):

    adjusted_width = get_column_max_width(column)
    column_letter = column[header_row].column_letter
    worksheet.column_dimensions[column_letter].width = adjusted_width


def format_column_to_currency(column):

    for cell in column:
        cell.number_format = '$#,##0.00'
            
def format_column_to_precentage(column):

    for cell in column:
        cell.number_format = '0.0%'

def add_borders_center_align_cell(cell):

    border_fmt = Border(top=Side(style='thin'),
                        right=Side(style='thin'),
                        bottom=Side(style='thin'),
                        left=Side(style='thin'))

    cell.border = border_fmt
    cell.alignment = Alignment(horizontal='center')


def add_borders_to_column(column, num_headers=1):

    for cell in column:
        add_borders_center_align_cell(cell)

    for row_num in range(0, num_headers):
        column[row_num].style = 'Pandas'
        


def conditional_formatting_redfill(worksheet,cell_range,operator,formula):

    red_font = Font(name='Calibri',
                   size=11,
                   bold=False,
                   italic=False,
                   vertAlign=None,
                   underline='none',
                   strike=False,
                   color='9C0006')
    red_fill = PatternFill(fill_type='solid',
                          start_color='FFC7CE',
                          end_color='FFC7CE')
    worksheet.conditional_formatting.add(cell_range,CellIsRule(operator=operator, formula=formula, stopIfTrue=False, font=red_font,fill=red_fill))
    
def conditional_formatting_greenfill(worksheet,cell_range,operator,formula):

    red_font = Font(name='Calibri',
                   size=11,
                   bold=False,
                   italic=False,
                   vertAlign=None,
                   underline='none',
                   strike=False,
                   color='006100')
    red_fill = PatternFill(fill_type='solid',
                          start_color='C6EFCE',
                          end_color='C6EFCE')
    worksheet.conditional_formatting.add(cell_range,CellIsRule(operator=operator, formula=formula, stopIfTrue=False, font=red_font,fill=red_fill))
